import sys

import yfinance as yf

import pandas as pd

from tqdm import tqdm

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


renaming_dictionary = {
    "call_lastPrice": "Último Precio (Call)",
    "put_lastPrice": "Último Precio (Put)",
    "call_change": "Cambio (Call)",
    "put_change": "Cambio (Put)",
    "put_percentChange": "Cambio % (Put)",
    "call_percentChange": "Cambio % (Call)",
    "call_bid": "Compra (Bid) Call",
    "put_bid": "Compra (Bid) Put",
    "call_ask": "Venta (Ask) Call",
    "put_ask": "Venta (Ask) Put",
    "call_volume": "Volumen (Call)",
    "put_volume": "Volumen (Put)",
    "call_openInterest": "Interés Abierto (Call)",
    "put_openInterest": "Interés Abierto (Put)",
    "strike": "Precio de Ejercicio",
    "call_inTheMoney": "",
    "put_inTheMoney": ""
}

def save_formated_excel(df: pd.DataFrame, file_name: str, option_name : str, price : int):
    file_name = file_name if file_name.endswith('.xlsx') else file_name + '.xlsx'

    # Save formatless file
    ruta = f"./{file_name}"
    df.to_excel(ruta, index=False, sheet_name=option_name)

    # apply format
    wb = load_workbook(ruta)
    ws = wb.active

    # styles
    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Format headers
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws[f"{col_letter}1"].font = header_font
        ws[f"{col_letter}1"].alignment = alignment
        ws[f"{col_letter}1"].border = thin_border

        # width auto
        max_length = max(len(str(ws[f"{col_letter}{row}"].value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)  # limit max width

    # Fills
    fill_itm_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Header color

    # Shift down 2 rows
    ws.insert_rows(1, amount=2)

    # Insert option name
    ws["B2"] = option_name
    ws["B2"].font = Font(bold=True, size=14)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    # Insert option price
    ws["C2"] = price
    ws["C2"].font = Font(bold=True, size=14)
    ws["C2"].number_format = '"$"#,##0.00'
    ws["C2"].alignment = Alignment(horizontal="left", vertical="center")

    # Aply format to headers
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        cell = ws[f"{col_letter}3"]
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border
        cell.fill = fill_header
        # Adjust column width
        max_length = max(len(str(ws[f"{col_letter}{row}"].value or "")) for row in range(3, ws.max_row + 1))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # Aply format to cells
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=ws.max_column):
        call_itm = row[0].value == "ITM"   # Columna A
        put_itm  = row[12].value == "ITM"  # Columna M
        for idx, cell in enumerate(row):
            cell.alignment = alignment
            cell.border = thin_border

            col_letter = get_column_letter(cell.column)

            # Strike
            if col_letter == "G":   
                cell.fill = fill_white

            # CALL ITM 
            elif call_itm and col_letter in ["A", "B", "C", "D", "E", "F"]:
                cell.fill = fill_itm_blue

            # PUT ITM 
            elif put_itm and col_letter in ["H", "I", "J", "K", "L", "M"]:
                cell.fill = fill_itm_blue

    wb.save(ruta)
    print(f"[✓] Archivo Excel guardado con formato: {ruta}")

def format_table(table):
        
        columns_to_exclude = ["call_inTheMoney", "put_inTheMoney"]
        columns_to_replace = [col for col in table.columns if col not in columns_to_exclude]
        table[columns_to_replace] = table[columns_to_replace].fillna("-")

        table["call_percentChange"] = table["call_percentChange"].apply(format_as_percentaje);
        table["put_percentChange"] = table["put_percentChange"].apply(format_as_percentaje);

        table["call_change"] = table["call_change"].apply(format_as_price);
        table["put_change"] = table["put_change"].apply(format_as_price);

        table["call_volume"] = table["call_volume"].apply(format_as_integer);
        table["put_volume"] = table["put_volume"].apply(format_as_integer);

        table["call_openInterest"] = table["call_openInterest"].apply(format_as_integer);
        table["put_openInterest"] = table["put_openInterest"].apply(format_as_integer);

        table["call_inTheMoney"] = table["call_inTheMoney"].apply(format_boolean);
        table["put_inTheMoney"] = table["put_inTheMoney"].apply(format_boolean);

def format_as_integer(value):
    try:
        if isinstance(value, str):
            if value.strip() == "-" or not value.strip():
                return "-"
            return int(value.replace('.', '').replace(',', ''))
        elif pd.isna(value):
            return "-"
        return int(value)
    except (ValueError, TypeError):
        return "-"

def format_boolean(value):
    return "ITM" if bool(value) else "OTM"

def format_as_percentaje(value):
    try:
        if value == "-" or pd.isna(value):
            return "-"
        return f"{float(value):.2f}%"
    except (TypeError, ValueError):
        return "-"

def format_as_price(value):
    try:
        if value == "-" or pd.isna(value):
            return "-"
        return f"{float(value):.2f}"
    except (TypeError, ValueError):
        return "-"
       
def fetch_options(symbol: str, expiration_date: str = None):
    try:
        ticker = yf.Ticker(symbol)
        expirations = ticker.options
        if expiration_date is None:
            expiration_date = expirations[0]

        data = ticker.history(period="1d")

        # Fetch last price
        precio = data['Close'].iloc[-1]

        options = ticker.option_chain(expiration_date)
        calls_df = options.calls
        puts_df = options.puts

        # Select usefull columns
        cols = ['lastPrice', 'change', 'percentChange','volume', 'openInterest', 'inTheMoney','strike']

        calls_df_sel = calls_df[cols].copy()
        puts_df_sel = puts_df[cols].copy()

        # Columns renaming
        calls_df_sel.columns = [f"call_{col}" if col != "strike" else "strike" for col in calls_df_sel.columns]
        puts_df_sel.columns = [f"put_{col}" if col != "strike" else "strike" for col in puts_df_sel.columns]

        # Merge by strike
        table = pd.merge(calls_df_sel, puts_df_sel, on='strike', how='outer');
        
        table["call_inTheMoney"] = table["call_inTheMoney"].astype("boolean")
        table["put_inTheMoney"] = table["put_inTheMoney"].astype("boolean")

        call_nulls = table["call_inTheMoney"].isna()
        put_nulls = table["put_inTheMoney"].isna()

        table.loc[call_nulls & ~put_nulls, "call_inTheMoney"] = ~table.loc[call_nulls & ~put_nulls, "put_inTheMoney"]
        table.loc[put_nulls & ~call_nulls, "put_inTheMoney"] = ~table.loc[put_nulls & ~call_nulls, "call_inTheMoney"]


        #Format the table
        format_table(table);
        
        #Order columns
        cols_order = (
            ['call_inTheMoney'] +
            ['call_lastPrice', 'call_change', 'call_percentChange', 'call_volume', 'call_openInterest'] +
            ['strike'] +
            ['put_lastPrice', 'put_change', 'put_percentChange', 'put_volume', 'put_openInterest'] +
            ['put_inTheMoney']
        )

        table = table[cols_order];

        # Save Straddle table
        table.rename(columns=renaming_dictionary, inplace=True)
        save_formated_excel(table, f"{symbol}_{expiration_date}", symbol, precio)

        return calls_df, puts_df, table

    except Exception as e:
        print(f"Error: {e}")
        return None, None, None

def load_from_txt(path_txt: str, expiration_date: str = None):
    try:
        with open(path_txt, 'r') as file:
            symbols = [line.strip() for line in file if line.strip()]

        print(f"\nTotal de símbolos a procesar: {len(symbols)}\n")

        for symbol in tqdm(symbols, desc="Procesando opciones", unit="símbolo"):
            _, _, tabla = fetch_options(symbol, expiration_date)
            if tabla is None:
                print(f"\n[!] No se pudo obtener la tabla para {symbol}")
            else:
                print(f"\n[✔] Archivo generado para {symbol}")

    except FileNotFoundError:
        print(f"No se encontró el archivo: {path_txt}")
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")

# ----------------- USO -----------------
if __name__ == "__main__":
    if len(sys.argv) > 1:
        path_txt = sys.argv[1]
    else:
        path_txt = input("No se detectó archivo. Ingresá el nombre del archivo .txt (con extensión): ").strip()

    if not path_txt.lower().endswith(".txt"):
        print("El archivo debe tener extensión .txt")
    else:
        fecha = None  # Podés cambiarlo por una fecha específica si querés
        load_from_txt(path_txt, fecha)
        input("Presioná ENTER para cerrar...")

